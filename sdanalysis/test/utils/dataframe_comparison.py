"""
dataframe_comparison.py - A module for comparing two pandas DataFrames with customizable behavior.
As opposed to pandas.testing.assert_frame_equal, this module is more flexible, capable of more than
just assertion.
"""

import os
import warnings
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def dataframe_differences(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    both_nan_equal: bool = False,
    tolerance: float = 1e-3,
) -> pd.DataFrame:
    """Given two dataframes with the same shape and columns, return a boolean dataframe of the
    same shape filled with True where the entries are equal,
    and False where they are not equal.
    If a position in both dataframes is np.NaN, the comparison for that cell is evaluated
    as "both_nan_equal"

    Parameters
    ----------
    df1 : pd.DataFrame
        first dataframe to compare
    df2 : pd.DataFrame
        second dataframe to compare
    both_nan_equal : bool, optional
        the value to fill cells of resulting boolean dataframe where
        both df1 and df2 contains np.NaN, by default False
    tolerance : float, optional
        the tolerance for float comparison, by default 1e-3

    Returns
    -------
    pd.DataFrame
        a boolean dataframe of the same shape as the input dataframes,
        filled with True where the entries are equal and not np.NaN,
        both_nan_equal where equal and np.NaN,
        and False where they are not equal.
    """

    if not (isinstance(df1, pd.DataFrame) and isinstance(df2, pd.DataFrame)):
        raise TypeError("Both inputs must be pandas DataFrames.")
    if not df1.shape == df2.shape:
        raise ValueError("Dataframes must have the same shape.")
    if not (df1.columns == df2.columns).all():
        raise ValueError("Dataframes must have the same columns.")

    # Initialize an empty DataFrame for the mask with the same shape
    mask = pd.DataFrame(True, index=df1.index, columns=df1.columns)

    # Loop through each column and compare values based on type
    for col in df1.columns:
        # Check for approximate equality for floats
        if np.issubdtype(df1[col].dtype, np.floating):
            mask[col] = np.isclose(df1[col], df2[col], atol=tolerance)
        else:  # Check for exact equality for other types
            mask[col] = df1[col] == df2[col]
        mask[col][df1[col].isna() & df2[col].isna()] = both_nan_equal
    return mask


def dataframes_equal(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    both_nan_equal: bool = False,
    tolerance: float = 1e-3,
) -> bool:
    """Compare two dataframes for equality. If a position in both dataframes is np.NaN,
    the comparison for that cell is evaluated as both_nan_equal.
    Parameters
    ----------
    df1 : pandas.DataFrame
        One dataframe to compare
    df2 : pandas.DataFrame
        Another dataframe to compare
    both_nan_equal : bool, optional
        Whether to consider np.NaN in both dataframes as equal, by default False
    tolerance : float, optional
        The tolerance for float comparison, by default 1e-3

    Returns
    -------
    bool
        True if the dataframes are equal, False otherwise
    """
    if not (isinstance(df1, pd.DataFrame) and isinstance(df2, pd.DataFrame)):
        raise TypeError("Both inputs must be pandas DataFrames.")
    if not df1.shape == df2.shape:
        return False
    if not (df1.columns == df2.columns).all():
        return False
    if not (df1.index == df2.index).all():
        return False
    comparison = dataframe_differences(
        df1, df2, both_nan_equal=both_nan_equal, tolerance=tolerance
    )
    return comparison.all().all()  # collapse over rows and then over columns


def series_equal(
    s1: pd.Series, s2: pd.Series, both_nan_equal: bool = False, tolerance: float = 0.001
) -> bool:
    """
    Compare two pandas Series for equality. If a position in both Series is np.NaN,
    the comparison for that cell is evaluated as both_nan_equal.

    Args:
        s1 (pd.Series): _description_
        s2 (pd.Series): _description_
    """
    if not (isinstance(s1, pd.Series) and isinstance(s2, pd.Series)):
        raise TypeError("Both inputs must be pandas DataFrames.")
    if not s1.shape == s2.shape:
        return False
    if not (s1.index == s2.index).all():
        return False
    if np.issubdtype(s1.dtype, np.floating):
        mask = np.isclose(s1, s2, atol=tolerance)
    else:
        mask = s1 == s2
    mask[s1.isna() & s2.isna()] = both_nan_equal
    return mask.all()


def write_diff_to_excel(df1: pd.DataFrame, df2: pd.DataFrame, fpath: str) -> None:
    """Given two dataframes, write a comparison of the two to an Excel file at the specified path.
    The comparison is color-coded: red for differences, green for similarities.
    Parameters
    ----------
    df1 : pd.DataFrame
        The first dataframe to compare
    df2 : pd.DataFrame
        The second dataframe to compare
    fpath : str
        The path to write the Excel file to (if it exists, it will not be overwritten)
    """
    if os.path.exists(fpath):
        warnings.warn(f"File already exists. Not overwriting:\n\t{fpath}")
        return
    highlight_diff = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )  # different values = red
    highlight_same = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )  # same values = green
    comparison = dataframe_differences(df1, df2)
    comparison.to_excel(fpath, index=False)
    workbook = load_workbook(fpath)
    worksheet = workbook.active
    for row in range(1, comparison.shape[0] + 1):  # rows (1-based index)
        # columns (1-based index)
        for col in range(1, comparison.shape[1] + 1):
            if comparison.iloc[row - 1, col - 1]:  # If there's a difference
                # row+1 to account for header
                worksheet.cell(row=row + 1, column=col).fill = highlight_same
            else:
                worksheet.cell(row=row + 1, column=col).fill = highlight_diff
    workbook.save(fpath)
